import io
import logging
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from app.services.case_ingestion import ingest_case
from fastapi.responses import StreamingResponse
from app.core.dependencies import get_lawyer, get_current_user
from app.core.database import supabase, supabase_admin
from app.core.email import send_invoice_email, send_payment_reminder_email

_log = logging.getLogger(__name__)
from pydantic import BaseModel, field_validator
from typing import Optional, List

from app.models.enums import InvoiceStatus
import secrets

router = APIRouter(prefix="/api/invoices", tags=["Billing"])

# ─── Helpers ────────────────────────────────────────────

def _auto_mark_overdue(firm_id: str) -> None:
    """Promote any PENDING invoice whose due_date has passed to OVERDUE."""
    today = date.today().isoformat()
    overdue_ids = (
        supabase.table("invoice")
        .select("id")
        .eq("firm_id", firm_id)
        .eq("status", "PENDING")
        .lt("due_date", today)
        .execute()
    )
    ids = [r["id"] for r in (overdue_ids.data or [])]
    if ids:
        supabase.table("invoice").update({"status": "OVERDUE"}).in_("id", ids).execute()

# ─── Schemas ────────────────────────────────────────────

class InvoiceItem(BaseModel):
    description: str
    quantity: float
    unit_price: float

class CreateInvoiceRequest(BaseModel):
    client_id: str
    case_id: Optional[str] = None
    items: List[InvoiceItem]
    tax_rate: float = 0
    due_date: date
    currency: str = "USD"
    notes: Optional[str] = None

    @field_validator("tax_rate")
    @classmethod
    def validate_tax_rate(cls, v):
        if not (0 <= v <= 20):
            raise ValueError("tax_rate must be between 0 and 20")
        return v

class UpdateInvoiceRequest(BaseModel):
    client_id: Optional[str] = None
    case_id: Optional[str] = None
    items: Optional[List[InvoiceItem]] = None
    tax_rate: Optional[float] = None
    due_date: Optional[date] = None
    currency: Optional[str] = None
    notes: Optional[str] = None

    @field_validator("tax_rate")
    @classmethod
    def validate_tax_rate(cls, v):
        if v is not None and not (0 <= v <= 20):
            raise ValueError("tax_rate must be between 0 and 20")
        return v

# ─── GET /api/invoices/analytics/summary ────────────────
# IMPORTANT: Must be declared BEFORE /{invoice_id} to avoid
# FastAPI interpreting "analytics" as an invoice_id path param.

def _get_lawyer_invoice_ids(firm_id: str, user_id: str) -> list:
    """Return invoice IDs visible to a regular lawyer (created by them or linked to their cases)."""
    team = supabase.table("case_team").select("case_id").eq("user_id", user_id).execute()
    lawyer_case_ids = [r["case_id"] for r in (team.data or [])]

    by_lawyer = {
        r["id"] for r in (
            supabase.table("invoice").select("id").eq("firm_id", firm_id).eq("lawyer_id", user_id).execute()
        ).data or []
    }
    by_case = set()
    if lawyer_case_ids:
        by_case = {
            r["id"] for r in (
                supabase.table("invoice").select("id").eq("firm_id", firm_id).in_("case_id", lawyer_case_ids).execute()
            ).data or []
        }
    return list(by_lawyer | by_case)


@router.get("/analytics/summary")
async def billing_analytics(current_user=Depends(get_lawyer)):
    firm_id  = current_user["firm_id"]
    is_admin = current_user["role"] in ("FIRM_ADMIN", "SUPER_ADMIN")

    _auto_mark_overdue(firm_id)

    inv_q = supabase.table("invoice").select("*").eq("firm_id", firm_id)
    if not is_admin:
        invoice_ids = _get_lawyer_invoice_ids(firm_id, current_user["id"])
        if not invoice_ids:
            return {"total_revenue": 0, "outstanding": 0, "overdue": 0, "total_invoices": 0, "collection_rate": 0}
        inv_q = inv_q.in_("id", invoice_ids)

    data = (inv_q.execute()).data or []

    total_revenue   = sum(i["total_amount"] for i in data if i["status"] == "PAID")
    outstanding     = sum(i["total_amount"] for i in data if i["status"] == "PENDING")
    overdue         = sum(i["total_amount"] for i in data if i["status"] == "OVERDUE")
    total_invoices  = len(data)
    paid_count      = len([i for i in data if i["status"] == "PAID"])
    collection_rate = round((paid_count / total_invoices * 100) if total_invoices > 0 else 0, 1)

    return {
        "total_revenue": total_revenue,
        "outstanding": outstanding,
        "overdue": overdue,
        "total_invoices": total_invoices,
        "collection_rate": collection_rate,
    }

# ─── GET /api/invoices/export ───────────────────────────
# Must be declared BEFORE /{invoice_id} to avoid routing conflict.

@router.get("/export")
async def export_invoices(
    format: str = "pdf",
    status: Optional[str] = None,
    current_user=Depends(get_lawyer),
):
    firm_id  = current_user["firm_id"]
    is_admin = current_user["role"] in ("FIRM_ADMIN", "SUPER_ADMIN")

    _auto_mark_overdue(firm_id)
    query = (
        supabase.table("invoice")
        .select("*, invoice_item(*), client(first_name, last_name, email)")
        .eq("firm_id", firm_id)
    )
    if not is_admin:
        invoice_ids = _get_lawyer_invoice_ids(firm_id, current_user["id"])
        if not invoice_ids:
            query = query.in_("id", [])
        else:
            query = query.in_("id", invoice_ids)
    if status:
        query = query.eq("status", status)
    invoices = (query.order("created_at", desc=True).execute()).data or []

    firm_res  = supabase.table("firm").select("name").eq("id", firm_id).single().execute()
    firm_name = (firm_res.data or {}).get("name", "LegalHub")

    if format.lower() == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=503, detail="openpyxl not installed. Run: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["Invoice #", "Client", "Status", "Issue Date", "Due Date", "Subtotal", "Tax", "Total", "Currency"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_idx, inv in enumerate(invoices, 2):
            client = inv.get("client") or {}
            client_name = f"{client.get('first_name','')} {client.get('last_name','')}".strip() or "—"
            ws.cell(row=row_idx, column=1, value=inv.get("invoice_number", ""))
            ws.cell(row=row_idx, column=2, value=client_name)
            ws.cell(row=row_idx, column=3, value=inv.get("status", ""))
            ws.cell(row=row_idx, column=4, value=str(inv.get("issue_date", "")))
            ws.cell(row=row_idx, column=5, value=str(inv.get("due_date", "")))
            ws.cell(row=row_idx, column=6, value=float(inv.get("subtotal", 0)))
            ws.cell(row=row_idx, column=7, value=float(inv.get("tax_amount", 0)))
            ws.cell(row=row_idx, column=8, value=float(inv.get("total_amount", 0)))
            ws.cell(row=row_idx, column=9, value=inv.get("currency", "USD"))
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="invoices_{date.today()}.xlsx"'},
        )

    # Default: PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        raise HTTPException(status_code=503, detail="reportlab not installed. Run: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f"{firm_name} — Invoice Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {date.today()} | Invoices: {len(invoices)}", styles["Normal"]))
    elements.append(Spacer(1, 16))

    col_headers = ["Invoice #", "Client", "Status", "Due Date", "Total", "Currency"]
    table_data  = [col_headers]
    for inv in invoices:
        client = inv.get("client") or {}
        client_name = f"{client.get('first_name','')} {client.get('last_name','')}".strip() or "—"
        table_data.append([
            inv.get("invoice_number", ""),
            client_name,
            inv.get("status", ""),
            str(inv.get("due_date", "")),
            f"{float(inv.get('total_amount', 0)):,.2f}",
            inv.get("currency", "USD"),
        ])

    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ALIGN",      (4, 0), (4, -1), "RIGHT"),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="invoices_{date.today()}.pdf"'},
    )


# ─── GET /api/invoices ──────────────────────────────────

@router.get("")
async def list_invoices(
    status: Optional[str] = None,
    client_id: Optional[str] = None,
    case_id: Optional[str] = None,
    current_user=Depends(get_current_user)
):
    firm_id  = current_user["firm_id"]
    is_admin = current_user["role"] in ("FIRM_ADMIN", "SUPER_ADMIN")

    _auto_mark_overdue(firm_id)

    query = (
        supabase.table("invoice")
        .select("*, invoice_item(*), client(id, first_name, last_name, email), case_file(id, title, case_number)")
        .eq("firm_id", firm_id)
    )

    if current_user["role"] == "CLIENT":
        client_result = (
            supabase.table("client")
            .select("id")
            .eq("user_id", current_user["id"])
            .single()
            .execute()
        )
        if not client_result.data:
            return []
        query = query.eq("client_id", client_result.data["id"])
    elif not is_admin:
        invoice_ids = _get_lawyer_invoice_ids(firm_id, current_user["id"])
        if not invoice_ids:
            return []
        query = query.in_("id", invoice_ids)

    if status:
        query = query.eq("status", status)
    if client_id:
        query = query.eq("client_id", client_id)
    if case_id:
        query = query.eq("case_id", case_id)

    result = query.order("created_at", desc=True).execute()
    return result.data

# ─── POST /api/invoices ─────────────────────────────────

@router.post("", status_code=201)
async def create_invoice(body: CreateInvoiceRequest, background_tasks: BackgroundTasks, current_user=Depends(get_lawyer)):
    subtotal   = sum(item.quantity * item.unit_price for item in body.items)
    tax_amount = subtotal * (body.tax_rate / 100)
    total      = subtotal + tax_amount

    invoice_number = f"INV-{secrets.token_hex(3).upper()}"

    data = {
        "firm_id":        current_user["firm_id"],
        "lawyer_id":      current_user["id"],
        "client_id":      body.client_id,
        "case_id":        body.case_id,
        "invoice_number": invoice_number,
        "status":         InvoiceStatus.DRAFT,
        "subtotal":       subtotal,
        "tax_rate":       body.tax_rate,
        "tax_amount":     tax_amount,
        "total_amount":   total,
        "currency":       body.currency,
        "issue_date":     date.today().isoformat(),
        "due_date":       body.due_date.isoformat(),
        "notes":          body.notes,
    }

    invoice = supabase.table("invoice").insert(data).execute()
    invoice_id = invoice.data[0]["id"]

    for item in body.items:
        supabase.table("invoice_item").insert({
            "invoice_id":  invoice_id,
            "description": item.description,
            "quantity":    item.quantity,
            "unit_price":  item.unit_price,
            "total":       item.quantity * item.unit_price,
        }).execute()

    supabase.table("notification").insert({
        "user_id": current_user["id"],
        "type":    "INVOICE_DUE",
        "title":   "Invoice Created",
        "message": f"{invoice_number} — {body.currency} {total:,.2f} due {body.due_date}.",
    }).execute()

    if body.case_id:
        background_tasks.add_task(ingest_case, body.case_id, current_user["firm_id"])
        supabase.table("case_timeline").insert({
            "case_id":      body.case_id,
            "firm_id":      current_user["firm_id"],
            "action":       f"Invoice created: {invoice_number} — {body.currency} {total:,.2f} due {body.due_date}",
            "performed_by": current_user["id"],
        }).execute()
    return invoice.data[0]

# ─── GET /api/invoices/:id ──────────────────────────────

@router.get("/{invoice_id}")
async def get_invoice(invoice_id: str, current_user=Depends(get_current_user)):
    result = (
        supabase.table("invoice")
        .select("*, invoice_item(*), client(id, first_name, last_name, email), case_file(id, title, case_number)")
        .eq("id", invoice_id)
        .eq("firm_id", current_user["firm_id"])
        .single()
        .execute()
    )
    if not result.data:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return result.data

# ─── PUT /api/invoices/:id ──────────────────────────────

@router.put("/{invoice_id}")
async def update_invoice(invoice_id: str, body: UpdateInvoiceRequest, background_tasks: BackgroundTasks, current_user=Depends(get_lawyer)):
    # Fetch current invoice to verify ownership
    existing = (
        supabase.table("invoice")
        .select("id, status, case_id")
        .eq("id", invoice_id)
        .eq("firm_id", current_user["firm_id"])
        .single()
        .execute()
    )
    if not existing.data:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if existing.data["status"] == "PAID":
        raise HTTPException(status_code=400, detail="Cannot edit a paid invoice")

    update_data: dict = {}

    if body.items is not None:
        subtotal   = sum(item.quantity * item.unit_price for item in body.items)
        tax_rate   = body.tax_rate if body.tax_rate is not None else 0
        tax_amount = subtotal * (tax_rate / 100)
        update_data.update({
            "subtotal":     subtotal,
            "tax_rate":     tax_rate,
            "tax_amount":   tax_amount,
            "total_amount": subtotal + tax_amount,
        })
        # Replace line items
        supabase.table("invoice_item").delete().eq("invoice_id", invoice_id).execute()
        for item in body.items:
            supabase.table("invoice_item").insert({
                "invoice_id":  invoice_id,
                "description": item.description,
                "quantity":    item.quantity,
                "unit_price":  item.unit_price,
                "total":       item.quantity * item.unit_price,
            }).execute()

    for field in ("client_id", "case_id", "currency", "notes"):
        val = getattr(body, field)
        if val is not None:
            update_data[field] = val
    if body.due_date is not None:
        update_data["due_date"] = body.due_date.isoformat()

    case_id = existing.data.get("case_id") or (update_data.get("case_id") if update_data else None)
    if update_data:
        result = (
            supabase.table("invoice")
            .update(update_data)
            .eq("id", invoice_id)
            .execute()
        )
        if case_id:
            background_tasks.add_task(ingest_case, case_id, current_user["firm_id"])
        return result.data[0]

    return existing.data

# ─── DELETE /api/invoices/:id ───────────────────────────

@router.delete("/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user=Depends(get_lawyer)):
    existing = (
        supabase.table("invoice")
        .select("id, status, invoice_number, case_id, total_amount, currency")
        .eq("id", invoice_id)
        .eq("firm_id", current_user["firm_id"])
        .single()
        .execute()
    )
    if not existing.data:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if existing.data["status"] == "PAID":
        raise HTTPException(status_code=400, detail="Cannot delete a paid invoice")
    inv = existing.data
    supabase.table("invoice_item").delete().eq("invoice_id", invoice_id).execute()
    supabase.table("invoice").delete().eq("id", invoice_id).execute()
    if inv.get("case_id"):
        supabase.table("case_timeline").insert({
            "case_id":      inv["case_id"],
            "firm_id":      current_user["firm_id"],
            "action":       f"Invoice deleted: {inv.get('invoice_number', invoice_id)} ({inv.get('currency', '')} {float(inv.get('total_amount', 0)):,.2f})",
            "performed_by": current_user["id"],
        }).execute()
    return {"message": "Invoice deleted"}

# ─── POST /api/invoices/:id/send ────────────────────────

@router.post("/{invoice_id}/send")
async def send_invoice(invoice_id: str, current_user=Depends(get_lawyer)):
    # Fetch invoice with items and client email in one query
    inv_res = (
        supabase.table("invoice")
        .select("*, invoice_item(*), client(id, first_name, last_name, email, user_id), case_file(id, title, case_number)")
        .eq("id", invoice_id)
        .eq("firm_id", current_user["firm_id"])
        .single()
        .execute()
    )
    if not inv_res.data:
        raise HTTPException(status_code=404, detail="Invoice not found")

    inv    = inv_res.data
    client = inv.get("client") or {}

    client_email = client.get("email")
    if not client_email:
        raise HTTPException(status_code=400, detail="Client has no email address on file")

    client_name = f"{client.get('first_name', '')} {client.get('last_name', '')}".strip() or "Client"

    # Fetch firm name
    firm_res  = supabase.table("firm").select("name").eq("id", current_user["firm_id"]).single().execute()
    firm_name = firm_res.data.get("name", "LegalHub") if firm_res.data else "LegalHub"

    # Send the email (raises on hard failure)
    lawyer_name = current_user.get("full_name") or current_user.get("email", "Your lawyer")

    try:
        send_invoice_email(
            to_email       = client_email,
            client_name    = client_name,
            firm_name      = firm_name,
            lawyer_name    = lawyer_name,
            invoice_number = inv["invoice_number"],
            issue_date     = str(inv.get("issue_date", "")),
            due_date       = str(inv.get("due_date",   "")),
            items          = inv.get("invoice_item", []),
            subtotal       = float(inv.get("subtotal",     0)),
            tax_rate       = float(inv.get("tax_rate",     0)),
            tax_amount     = float(inv.get("tax_amount",   0)),
            total_amount   = float(inv.get("total_amount", 0)),
            currency       = inv.get("currency", "USD"),
            notes          = inv.get("notes") or "",
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Email delivery failed: {e}")

    # Mark as PENDING only after successful send
    supabase.table("invoice").update({
        "status": InvoiceStatus.PENDING
    }).eq("id", invoice_id).execute()

    supabase.table("notification").insert({
        "user_id": current_user["id"],
        "type":    "INVOICE_DUE",
        "title":   "Invoice Sent",
        "message": f"{inv['invoice_number']} sent to {client_name} — due {inv.get('due_date', '')}.",
    }).execute()

    client_user_id = client.get("user_id")
    _log.info(
        f"[send_invoice] client_id={client.get('id')} "
        f"client_user_id={client_user_id!r} "
        f"invoice={inv['invoice_number']}"
    )
    if not client_user_id:
        _log.warning(
            f"[send_invoice] ⚠️  No user_id on client {client.get('id')} — "
            "client has not accepted portal invitation yet. Notification skipped."
        )
    else:
        try:
            result = supabase_admin.table("notification").insert({
                "user_id": client_user_id,
                "type":    "INVOICE_DUE",
                "title":   "New Invoice from Your Attorney",
                "message": f"Invoice {inv['invoice_number']} — {inv.get('currency', 'USD')} {float(inv.get('total_amount', 0)):,.2f} · Due {inv.get('due_date', '')}.",
            }).execute()
            _log.info(
                f"[send_invoice] ✅ Client notification inserted — "
                f"notif_id={result.data[0].get('id') if result.data else 'unknown'}"
            )
        except Exception as e:
            _log.error(f"[send_invoice] ❌ Client notification failed: {e}")

    if inv.get("case_id"):
        supabase.table("case_timeline").insert({
            "case_id":      inv["case_id"],
            "firm_id":      current_user["firm_id"],
            "action":       f"Invoice sent: {inv['invoice_number']} to {client_name} — {inv.get('currency', 'USD')} {float(inv.get('total_amount', 0)):,.2f} due {inv.get('due_date', '')}",
            "performed_by": current_user["id"],
        }).execute()

    return {"message": f"Invoice sent to {client_email}"}

# ─── POST /api/invoices/:id/cancel ─────────────────────

@router.post("/{invoice_id}/cancel")
async def cancel_invoice(invoice_id: str, current_user=Depends(get_lawyer)):
    existing = (
        supabase.table("invoice")
        .select("id, status")
        .eq("id", invoice_id)
        .eq("firm_id", current_user["firm_id"])
        .single()
        .execute()
    )
    if not existing.data:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if existing.data["status"] != "DRAFT":
        raise HTTPException(status_code=400, detail="Only DRAFT invoices can be cancelled")
    supabase.table("invoice").update({"status": InvoiceStatus.CANCELLED}).eq("id", invoice_id).execute()
    return {"message": "Invoice cancelled"}

# ─── POST /api/invoices/:id/reminder ────────────────────

@router.post("/{invoice_id}/reminder")
async def send_reminder(invoice_id: str, current_user=Depends(get_lawyer)):
    inv_res = (
        supabase.table("invoice")
        .select("*, client(id, first_name, last_name, email)")
        .eq("id", invoice_id)
        .eq("firm_id", current_user["firm_id"])
        .single()
        .execute()
    )
    if not inv_res.data:
        raise HTTPException(status_code=404, detail="Invoice not found")

    inv    = inv_res.data
    client = inv.get("client") or {}

    client_email = client.get("email")
    if not client_email:
        raise HTTPException(status_code=400, detail="Client has no email address on file")

    client_name = f"{client.get('first_name', '')} {client.get('last_name', '')}".strip() or "Client"
    firm_res    = supabase.table("firm").select("name").eq("id", current_user["firm_id"]).single().execute()
    firm_name   = firm_res.data.get("name", "LegalHub") if firm_res.data else "LegalHub"
    lawyer_name = current_user.get("full_name") or current_user.get("email", "Your lawyer")

    due_date_str = str(inv.get("due_date", "")) if inv.get("due_date") else "—"

    try:
        send_payment_reminder_email(
            to_email       = client_email,
            client_name    = client_name,
            firm_name      = firm_name,
            lawyer_name    = lawyer_name,
            invoice_number = inv["invoice_number"],
            due_date       = due_date_str,
            total_amount   = float(inv.get("total_amount", 0)),
            currency       = inv.get("currency", "USD"),
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Email delivery failed: {e}")

    if inv.get("case_id"):
        supabase.table("case_timeline").insert({
            "case_id":      inv["case_id"],
            "firm_id":      current_user["firm_id"],
            "action":       f"Payment reminder sent to {client_name}: {inv['invoice_number']} — {inv.get('currency', 'USD')} {float(inv.get('total_amount', 0)):,.2f} due {inv.get('due_date', '')}",
            "performed_by": current_user["id"],
        }).execute()

    return {"message": f"Payment reminder sent to {client_email}"}
